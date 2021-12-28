# created by kbkozlev 28.12.2021
import time
from pynput.keyboard import Key, Controller as KeyController
from pynput.mouse import Controller as MouseController
from openpyxl import load_workbook

Excel = input("Import Excel file here : ")
wb = load_workbook(Excel)  # Work Book
ws = wb['Sheet1']  # Work Sheet
shipments = ws['A']  # Shipments
shipments_list = [shipments[x].value for x in range(1, len(shipments))]

key = KeyController()
mouse = MouseController()


def f8_toggle():
    key.press(Key.f8)
    key.release(Key.f8)


def tab_toggle():
    key.press(Key.tab)
    key.release(Key.tab)


def esc_toggle():
    key.press(Key.esc)
    key.release(Key.esc)


def del_toggle():
    key.press(Key.delete)
    key.release(Key.delete)


minutes = (((len(shipments_list) * 2) + (len(shipments_list) // 10) * 5) // 60)

xx = 5
wait = 0

print(str(len(shipments_list)) + " Shipments imported.\n")
print("This task will take ~ " + str(minutes) + " min to complete\n")

input("Press any key to start:\n")
while xx > 0:
    print("Starting in " + str(xx) + " seconds")
    xx -= 1
    time.sleep(1)

for x in range(len(shipments_list)):
    start_position = mouse.position

    tab_toggle()
    del_toggle()
    key.type(str(shipments_list[x]))
    time.sleep(0.5)
    f8_toggle()
    time.sleep(0.5)
    esc_toggle()
    time.sleep(1)
    wait += 1

    end_position = mouse.position

    if start_position != end_position:
        break
    elif wait == 10:
        print("Overload pause 5 sec.")
        time.sleep(5)
        wait = 0

